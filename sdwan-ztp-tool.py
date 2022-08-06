import time

import eel
from tkinter import filedialog
from tkinter import *
from tkinter import *
from openpyxl import load_workbook
import json, requests, ipaddress, sys, platform, io

requests.packages.urllib3.disable_warnings()

# Set web files folder and optionally specify which file types to check for eel.expose()
#   *Default allowed_extensions are: ['.js', '.html', '.txt', '.htm', '.xhtml']
eel.init('web', allowed_extensions=['.js', '.html'])


def sendupdate(return_html):
    eel.pageupdate(return_html)


### Export ADOM Functions

def export_adom(adomname):
    global export_info
    global sdwan_template_list
    global polpkg_list
    export_info = {"vars": ["adom_name", "adom_desc"], "settings": []}
    requestid = 1

    ## Get ADOM Info
    newdict = {"url": "/dvmdb/adom/", "method": "add", "data": []}

    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/dvmdb/adom/" + adomname
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    parsed = json.loads(res.text)
    # print(json.dumps(parsed, indent=4, sort_keys=True))

    dvmdbadom = parsed['result'][0]['data']
    dvmdbadom.pop('uuid', None)
    dvmdbadom.pop('oid', None)
    dvmdbadom["name"] = "$(adom_name)"
    dvmdbadom["desc"] = "$(adom_desc)"
    if adomname == "root":
        dvmdbadom["flags"] = 2312

    newdict["data"].append(dvmdbadom)
    export_info["settings"].append(newdict)

    ## standard objects (objects which can be exported and imported with out changing anything, mutliple objects can be created in a list)

    sdwan_template_list = []
    polpkg_list = []

    std_objects = {
        "clitemplate": ["/pm/config/adom/$(adom_name)/obj/cli/template",
                        "/pm/config/adom/" + adomname + "/obj/cli/template", ["modification-time"]],
        "clitemplate-group": ["/pm/config/adom/$(adom_name)/obj/cli/template-group",
                              "/pm/config/adom/" + adomname + "/obj/cli/template-group", ["modification-time"]],
        "sdwaninterface": ["/pm/config/adom/$(adom_name)/obj/dynamic/virtual-wan-link/members/",
                           "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/",
                           ["dynamic_mapping", "obj seq"]],
        "sdwanservers": ["/pm/config/adom/$(adom_name)/obj/dynamic/virtual-wan-link/server/",
                         "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/server/",
                         ["dynamic_mapping"]],
        "sdwantemplates": ["/pm/wanprof/adom/$(adom_name)",
                           "/pm/wanprof/adom/" + adomname,
                           ["scope member", "oid"]],
        "addrobjs": ["pm/config/adom/$(adom_name)/obj/firewall/address/",
                     "pm/config/adom/" + adomname + "/obj/firewall/address/",
                     ["uuid", "dynamic_mapping"]],
        "addrobjs_grp": ["pm/config/adom/$(adom_name)/obj/firewall/addrgrp/",
                         "pm/config/adom/" + adomname + "/obj/firewall/addrgrp/",
                         ["uuid", "dynamic_mapping"]],
        "intfobjs": ["pm/config/adom/$(adom_name)/obj/dynamic/interface/",
                     "pm/config/adom/" + adomname + "/obj/dynamic/interface/",
                     ["uuid", "dynamic_mapping"]],
        "applist": ["pm/config/adom/$(adom_name)/obj/application/list",
                    "pm/config/adom/" + adomname + "/obj/application/list",
                    ["uuid", "dynamic_mapping", "obj seq"]],
        "appgrp": ["pm/config/adom/$(adom_name)/obj/application/group",
                   "pm/config/adom/" + adomname + "/obj/application/group",
                   ["uuid", "dynamic_mapping", "obj seq"]],
        "service": ["pm/config/adom/$(adom_name)/obj/firewall/service/custom",
                    "pm/config/adom/" + adomname + "/obj/firewall/service/custom",
                    ["uuid", "dynamic_mapping", "obj seq"]],
        "servicegrp": ["pm/config/adom/$(adom_name)/obj/firewall/service/group",
                       "pm/config/adom/" + adomname + "/obj/firewall/service/group",
                       ["uuid", "dynamic_mapping", "obj seq"]],
        "polpkg": ["pm/pkg/adom/$(adom_name)",
                   "pm/pkg/adom/" + adomname,
                   ["scope member", "oid"]]

    }

    for objecturls in std_objects:
        get_and_add(std_objects, objecturls)

    ## Get SDWAN Template Details
    for sdwan_template in sdwan_template_list:
        get_and_add({"sdwan_member": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/member",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/member",
            ["obj seq"]]}, "sdwan_member")

        get_and_add({"sdwan_hlth": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/health-check",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/health-check",
            ["obj seq"]]}, "sdwan_hlth")

        get_and_add({"sdwan_service": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/service",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/service",
            ["obj seq"]]}, "sdwan_service")

    ## Get Policy Package Details
    for polpkg in polpkg_list:
        get_and_add({"polpkg_policy": ["pm/config/adom/$(adom_name)/pkg/" + polpkg + "/firewall/policy",
                                       "pm/config/adom/" + adomname + "/pkg/" + polpkg + "/firewall/policy",
                                       ["obj seq", "_policy_block"]]}, "polpkg_policy")

    # print(json.dumps(export_info, indent=4, sort_keys=True))
    return json.dumps(export_info, indent=4, sort_keys=True)


def get_and_add(std_objects, objecturls):
    newdict = {"url": std_objects[objecturls][0], "method": "add", "data": []}
    if objecturls == "sdw_members":
        newdict = {"url": std_objects[objecturls][0], "method": "replace", "data": []}
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": std_objects[objecturls][1]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    parsed = json.loads(res.text)
    # print(json.dumps(parsed, indent=4, sort_keys=True))

    newdata = parsed['result'][0]['data']
    for index, config in enumerate(newdata):
        for popitem in std_objects[objecturls][2]:
            if popitem in newdata[index].keys():
                newdata[index].pop(popitem)

        ignore_addr_obj = ["wildcard.dropbox.com", "wildcard.google.com", "SSLVPN_TUNNEL_ADDR1", "all", "gmail.com",
                           "login.microsoft.com", "login.microsoftonline.com", "login.windows.net", "none",
                           "FABRIC_DEVICE", "FIREWALL_AUTH_PORTAL_ADDRESS"]
        if objecturls == "addrobjs":
            if newdata[index]['name'] in ignore_addr_obj:
                newdata[index] = {}

        ignore_addrgrp_obj = ["G Suite", "Microsoft Office 365"]
        if objecturls == "addrobjs_grp":
            if newdata[index]['name'] in ignore_addrgrp_obj:
                newdata[index] = {}

        ignore_sdwan_hlth = ["Default_AWS", "Default_FortiGuard", "Default_Gmail", "Default_Google Search",
                             "Default_Office_365"]
        if objecturls == "sdwan_hlth":
            if newdata[index]['name'] in ignore_sdwan_hlth:
                newdata[index] = {}

        if objecturls == "sdwantemplates":
            sdwan_template_list.append(newdata[index]['name'])

        if objecturls == "polpkg":
            polpkg_list.append(newdata[index]['name'])

        if objecturls == "sdwan_service":
            if isinstance(newdata[index]['sla'], list):
                for index2, config2 in enumerate(newdata[index]['sla']):
                    if "obj seq" in newdata[index]['sla'][index2].keys():
                        newdata[index]['sla'][index2].pop("obj seq")

        ignore_applist_obj = ["block-high-risk","default", "sniffer-profile", "wifi-default"]
        if objecturls == "applist":
            if isinstance(newdata[index]['entries'], list):
                for index2, config2 in enumerate(newdata[index]['entries']):
                    if "obj seq" in newdata[index]['entries'][index2].keys():
                        newdata[index]['entries'][index2].pop("obj seq")

            if newdata[index]['name'] in ignore_applist_obj:
                newdata[index] = {}

        ignore_service_obj = ["ALL", "ALL_TCP", "ALL_UDP", "ALL_ICMP", "ALL_ICMP6", "GRE", "GTP", "AH", "ESP", "AOL",
                              "BGP", "DHCP", "DNS", "FINGER", "FTP", "FTP_GET", "FTP_PUT", "GOPHER", "H323", "HTTP",
                              "HTTPS", "IKE", "IMAP", "IMAPS", "Internet-Locator-Service", "IRC", "L2TP", "LDAP",
                              "NetMeeting", "NFS", "NNTP", "NTP", "OSPF", "PC-Anywhere", "PING", "TIMESTAMP",
                              "INFO_REQUEST", "INFO_ADDRESS", "ONC-RPC", "DCE-RPC", "POP3", "POP3S", "PPTP", "QUAKE",
                              "RAUDIO", "REXEC", "RIP", "RLOGIN", "RSH", "SCCP", "SIP", "SIP-MSNmessenger", "SAMBA",
                              "SMTP", "SMTPS", "SNMP", "SSH", "SYSLOG", "TALK", "TELNET", "TFTP", "MGCP", "UUCP",
                              "VDOLIVE", "WAIS", "WINFRAME", "X-WINDOWS", "PING6", "MS-SQL", "MYSQL", "RDP", "VNC",
                              "DHCP6", "SQUID", "SOCKS", "WINS", "RADIUS", "RADIUS-OLD", "CVSPSERVER", "AFS3",
                              "TRACEROUTE", "RTSP", "MMS", "KERBEROS", "LDAP_UDP", "SMB", "NONE", "webproxy"]
        if objecturls == "service":
            if newdata[index]['name'] in ignore_service_obj:
                newdata[index] = {}

        ignore_servicegrp_obj = ["Email Access",
                                 "Web Access", "Windows AD", "Exchange Server"]
        if objecturls == "servicegrp":
            if newdata[index]['name'] in ignore_servicegrp_obj:
                newdata[index] = {}

    while {} in newdata:
        newdata.remove({})
    newdict["data"] = newdata
    export_info["settings"].append(newdict)

### Start copy from draft

def openbook(filename):
    headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data, device_daddr6_data, device_vpn_data = "", "", "", "", "", "", ""
    try:
        with open(filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(in_mem_file, data_only=True)

        try:
            ws = wb['Devices']
        except:
            try:
                ws = wb['Sheet1']
            except:
                ws = wb.active

        print("cell A1 value = " + ws.cell(1, 1).value)

        if ws.cell(1, 1).value == "Device_Name":
            print("ok")
            ## Get Columns
            headings = ['nul']
            col = 0
            blankheading = 0
            while blankheading < 3:
                col += 1
                if ws.cell(row=1, column=col).value == None:
                    blankheading += 1
                else:
                    headings.append(ws.cell(row=1, column=col).value)

            ## Get all Device Rows
            AllDevicesList = []
            device_meta_data = {}
            device_dint_data = {}
            device_sdwanint_data = {}
            device_daddr_data = {}
            device_vpn_data = {}
            device_daddr6_data = {}
            blankrow = 0
            row = 1

            while blankrow < 3:
                row += 1
                if ws.cell(row=row, column=1).value is None:
                    blankrow += 1
                else:
                    # get device detail in row

                    col = 1
                    newdict = {}
                    for i in headings:
                        if i != 'nul':
                            if ws.cell(row=row, column=col).value is None:
                                newdict[i] = ""
                            else:
                                newdict[i] = str(ws.cell(row=row, column=col).value)
                            if i == "Device_Name":
                                device_meta_data[newdict['Device_Name']] = {}
                                device_meta_data[newdict['Device_Name']]['Device_Name'] = newdict['Device_Name']
                                device_dint_data[newdict['Device_Name']] = {}
                                device_sdwanint_data[newdict['Device_Name']] = {}
                                device_daddr_data[newdict['Device_Name']] = {}
                                device_vpn_data[newdict['Device_Name']] = {}
                                device_daddr6_data[newdict['Device_Name']] = {}
                            if i == "Device_SN":
                                device_meta_data[newdict['Device_Name']]['Device_SN'] = newdict['Device_SN']
                            if i[0:5] == "meta_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_meta_data[newdict['Device_Name']][i[5:]] = ""
                                else:
                                    device_meta_data[newdict['Device_Name']][i[5:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:5] == "dint_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_dint_data[newdict['Device_Name']][i[5:]] = ""
                                else:
                                    device_dint_data[newdict['Device_Name']][i[5:]] = str(
                                        ws.cell(row=row, column=col).value).split(",")
                            if i[0:9] == "sdwanint_":
                                sdwanintsettings = i[9:].split("|")
                                try:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]]
                                except:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]] = {}
                                if ws.cell(row=row, column=col).value is not None:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]][
                                        sdwanintsettings[1]] = str(ws.cell(row=row, column=col).value)

                            if i[0:6] == "daddr_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_daddr_data[newdict['Device_Name']][i[6:]] = ""
                                else:
                                    device_daddr_data[newdict['Device_Name']][i[6:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:7] == "daddr6_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_daddr6_data[newdict['Device_Name']][i[7:]] = ""
                                else:
                                    device_daddr6_data[newdict['Device_Name']][i[7:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:6] == "vpn_OL":
                                if ws.cell(row=row, column=col).value is None:
                                    device_vpn_data[newdict['Device_Name']][i[4:]] = ""
                                else:
                                    device_vpn_data[newdict['Device_Name']][i[4:]] = str(
                                        ws.cell(row=row, column=col).value)

                            col += 1

                    AllDevicesList.append(newdict)
        else:
            AllDevicesList = "worksheet"

    except Exception as e:
        AllDevicesList = "workbook"
        print(e)

    wb = None
    return AllDevicesList, headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data, device_daddr6_data, device_vpn_data


def get_workspace():
    if fmgurl.find("fortimanager.forticloud.com") != -1:
        print("### This is FortiManager Cloud - Workspacemode not supported")
        workspacemode = 0
    else:
        requestid = 1
        jsondata = {
            "method": "get",
            "params": [
                {
                    "url": "/cli/global/system/global"
                }
            ],
            "id": requestid,
            "session": fmg_sessionid
        }

        # print("Request:")
        # print(json.dumps(jsondata, indent=4, sort_keys=True))
        res = session.post(fmgurl, json=jsondata, verify=False)
        response = json.loads(res.text)
        # print("Response:")
        # print(json.dumps(response, indent=4, sort_keys=True))

        try:
            workspacemode = response['result'][0]['data']['workspace-mode']
        except:
            workspacemode = 3
    return workspacemode


def lock_adom(adom):
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/lock",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def unlock_adom(adom):
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/unlock",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def workspace_commit(adom):
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/devprof/adom/" + adom + "/default",
                "data": {
                    "description": str(time.time())
                }
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))



    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/commit",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def get_meta():
    requestid = 2
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    meta_json = json.loads(res.text)
    return meta_json['result'][0]['data']


def create_meta(newname):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
                "data": {
                    "importance": 0,
                    "length": 255,
                    "name": newname,
                    "status": 1
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    try:
        metacreate_data = json.loads(res.text)
        ret_meta = "Meta Field " + newname + " created.<br>\n"
    except:
        ret_meta = "Failed to create Meta Field " + newname
    return ret_meta


def track_model_task(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)

    return ret_status


def add_model_device(adomname, devicename, sn, platform, prefer_img, fmg_adom_osver, fmg_adom_mr):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvm/cmd/add/device",
                "data": {
                    "adom": adomname,
                    "flags": [
                        "create_task",
                        "nonblocking"
                    ],
                    "device": {
                        "name": devicename,
                        "adm_usr": "admin",
                        "adm_pass": "",
                        "platform_str": platform,
                        "prefer_img_ver": prefer_img,
                        "mgmt_mode": 3,
                        "flags": 67371040,
                        "sn": sn,
                        "os_ver": fmg_adom_osver,
                        "mr": fmg_adom_mr
                    }
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['taskid'])
    return last_task


def update_device(adom, devicename):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "meta fields": device_meta_data[devicename],
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)


def add_device_coords(devicename, adom, long, lat):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "longitude": long,
                    "latitude": lat
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)
    json_devcoords = json.loads(res.text)
    status_devcoords = json_devcoords['result'][0]['status']['message']
    return status_devcoords


def change_admpass(devicename, adom, newpass):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "adm_pass": newpass
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("Change admin password: " + res.text)
    json_admpw = json.loads(res.text)
    status_admpw = json_admpw['result'][0]['status']['message']
    return status_admpw


def assign_cli_template(adom, template, devicename):
    ## template or template group
    template_string = "template"

    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/template-group/" + template
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_res = json.loads(res.text)
    if json_res['result'][0]['status']['message'] == "OK":
        template_string = "template-group"

    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/" + template_string + "/" + template + "/scope member",
                "data":
                    [
                        {
                            "name": devicename,
                            "vdom": "root"
                        }
                    ]

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    # print("Request:")
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_assignclitemplate = json.loads(res.text)
    # print("Response:")
    # print(json.dumps(json_assignclitemplate, indent=4, sort_keys=True))
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def unassign_cli_template(adom, template, devicename):
    ## template or template group
    template_string = "template"

    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/template-group/" + template
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_res = json.loads(res.text)
    if json_res['result'][0]['status']['message'] == "OK":
        template_string = "template-group"

    requestid = 1
    jsondata = {
        "method": "delete",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/" + template_string + "/" + template + "/scope member",
                "data":
                    [
                        {
                            "name": devicename,
                            "vdom": "root"
                        }
                    ]

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def quickinstall(adom, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "/securityconsole/install/device",
                "data": {
                    "adom": adom,
                    "scope": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ]
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_quickinstall = json.loads(res.text)
    taskid_qi = json_quickinstall['result'][0]['data']['task']
    return taskid_qi


def track_quickinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status


def add_install_target(device, adomname, vdomname, pkg):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "pm/pkg/adom/" + adomname + "/" + pkg + "/scope member",
                "data": [
                    {
                        "name": device,
                        "vdom": vdomname
                    },
                ]
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assignppkg = json.loads(res.text)
    status_ppkg = json_assignppkg['result'][0]['status']['message']
    return status_ppkg


def add_device_to_group(device, adomname, vdomname, groupname):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "dvmdb/adom/" + adomname + "/group/" + groupname + "/object member",
                "data": [
                    {
                        "name": device,
                        "vdom": vdomname
                    },
                ]
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_devgroup = json.loads(res.text)
    status_devgroup = json_devgroup['result'][0]['status']['message']
    return status_devgroup


def install_pkg(pkg, adomname, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "securityconsole/install/package",
                "data":
                    {"pkg": pkg,
                     "adom": adomname,
                     "scope": [
                         {
                             "name": devicename,
                             "vdom": vdom
                         }
                     ]
                     }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print()
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['task'])
    return last_task


def track_policyinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status


def add_policy_interface_member(adomname, newinterfacename, realinterface, devicename):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/dynamic/interface/" + newinterfacename + "/dynamic_mapping",
                "data":
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": "root"
                            }
                        ],
                        "local-intf": realinterface,
                        "intrazone-deny": 0
                    }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    # print("Request:")
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_mapdint = json.loads(res.text)
    # print("Response:")
    # print(json.dumps(json_mapdint, indent=4, sort_keys=True))
    status_mapdint = json_mapdint['result'][0]['status']['message']
    return status_mapdint


def add_sdwaninterface_mapping(adomname, devicename, interfacename, vdom):
    ## get settings for base SDWAN interface
    requestid = 1

    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename,
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_sdwanint_res = json.loads(res.text)
    if json_sdwanint_res['result'][0]['status']['message'] == "OK":
        json_sdwanint = json_sdwanint_res['result'][0]['data']
        json_sdwanint.pop('dynamic_mapping', None)
        json_sdwanint.pop('obj seq', None)
        json_sdwanint.pop('name', None)

        json_sdwanint["_scope"] = [
            {
                "name": devicename,
                "vdom": vdom
            }
        ]

        proceed_makesdwanint = 0
        for key in device_sdwanint_data[devicename][interfacename]:
            proceed_makesdwanint = 1
            json_sdwanint[key] = device_sdwanint_data[devicename][interfacename][key]

        if proceed_makesdwanint == 1:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename + "/dynamic_mapping",
                        "data": json_sdwanint
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            print(json_sdwanint)
            print(res.text)
            json_mapsdwanint = json.loads(res.text)
            status_mapsdwanint = json_mapsdwanint['result'][0]['status']['message']
        else:
            status_mapsdwanint = "NoData"
    else:
        status_mapsdwanint = json_sdwanint_res['result'][0]['status']['message']
    return status_mapsdwanint


def assign_sdwan_template(adom, sdwantemplate, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/pm/wanprof/adom/" + adom + "/" + sdwantemplate + "/scope member",
                "data": [
                    {
                        "name": devicename,
                        "vdom": vdom
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def add_daddr(adomname, daddrobj, newaddr, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)

    current_int_result = json.loads(res.text)
    if current_int_result['result'][0]['status']['message'] == "OK":
        current_int = current_int_result['result'][0]['data']

        result_msg = "unknown error"

        submit = False

        if current_int['type'] == 0:
            try:
                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "allow-routing": current_int['allow-routing'],
                        "subnet": [
                            str(ipaddress.ip_network(newaddr).network_address),
                            str(ipaddress.ip_network(newaddr).netmask)
                        ],
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not decode ip address into network_address/netmask"
        elif current_int['type'] == 1:
            try:
                newaddr.strip(" ")
                splitaddr = newaddr.split("-")

                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "allow-routing": current_int['allow-routing'],
                        "end-ip": splitaddr[1].strip(),
                        "start-ip": splitaddr[0].strip()
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not calculate IP RANGE"

        if submit is True:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj + "/dynamic_mapping",
                        "data": addrsettings
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            # print(res.text)
            json_result = json.loads(res.text)
            result_msg = json_result['result'][0]['status']['message']
    else:
        result_msg = current_int_result['result'][0]['status']['message']
    return result_msg


def add_daddr6(adomname, daddrobj, newaddr, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/firewall/address6/" + daddrobj
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)

    current_int_result = json.loads(res.text)
    if current_int_result['result'][0]['status']['message'] == "OK":
        current_int = current_int_result['result'][0]['data']

        result_msg = "unknown error"

        submit = False

        if current_int['type'] == 0:
            try:
                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "ip6": newaddr
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not decode ip address into network_address/netmask"
        elif current_int['type'] == 1:
            try:
                newaddr.strip(" ")
                splitaddr = newaddr.split("-")

                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "end-ip": splitaddr[1].strip(),
                        "start-ip": splitaddr[0].strip()
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not calculate IP RANGE"

        if submit is True:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/firewall/address6/" + daddrobj + "/dynamic_mapping",
                        "data": addrsettings
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            # print(res.text)
            json_result = json.loads(res.text)
            result_msg = json_result['result'][0]['status']['message']
    else:
        result_msg = current_int_result['result'][0]['status']['message']
    return result_msg


def add_cert_template(device, adom, cert):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "/securityconsole/sign/certificate/template",
                "data": [
                    {
                        "adom": adom,
                        "scope": [
                            {
                                "name": device,
                                "vdom": ""
                            }
                        ],
                        "template": "adom/" + adom + "/obj/certificate/template/" + cert
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assigncerttemplate = json.loads(res.text)
    status_assigncerttemplate = json_assigncerttemplate['result'][0]['status']['message']
    return status_assigncerttemplate


def add_vpn_overlay(adom, overlayname, authpasswd):
    # Adds a VPN Community to FortiManager
    # @darryl
    requestid = 1
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/config/adom/" + adom + "/obj/vpnmgr/vpntable",
                "data": [
                    {
                        "name": overlayname,
                        "description": "Overlay Created by ZTP Tool",
                        "topology": 2,
                        "psk-auto-generate": "enable",
                        "ike1keylifesec": 28800,
                        "ike1dpd": 1,
                        "ike1natkeepalive": 10,
                        "ike2keylifesec": 1800,
                        "ike2keylifekbs": 5120,
                        "ike2keepalive": 1,
                        "intf-mode": 0,
                        "fcc-enforcement": 0,
                        "ike-version": 1,
                        "negotiate-timeout": 30,
                        "inter-vdom": 0,
                        "auto-zone-policy": 0,
                        "npu-offload": 1,
                        "authmethod": 1,
                        "ike1dhgroup": 12,
                        "dpd": 3,
                        "localid-type": 0,
                        "ike1mode": 1,
                        "ike1nattraversal": 1,
                        "ike1proposal": [
                            "aes128-sha256",
                            "aes256-sha256"
                        ],
                        "ike2autonego": 0,
                        "ike2dhgroup": 12,
                        "ike2keylifetype": 1,
                        "pfs": 1,
                        "ike2proposal": [
                            "aes128-sha256",
                            "aes256-sha256"
                        ],
                        "replay": 1
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("### add_vpn_overlay ")
    json_add_vpn_overlay = json.loads(res.text)
    status_add_vpn_overlay = json_add_vpn_overlay['result'][0]['status']['message']
    return status_add_vpn_overlay


def add_vpn_hub(adom, overlayname, interface, authpasswd, devicename, vdom, oNetwork):
    # Adds a hub to an Existing VPN community in FortiManager

    # need to add Check Overlay Exists/Check Community Exists?
    # @Darryl
    # Enhancement - Need to update Exiting Overlay\Node ID number, otherwise use a new ID.
    # Note - this currently uses ID 0 - which means next available ID number - if this imports twice you will get two entries

    oTemp = ipaddress.ip_network(oNetwork, strict=False)

    requestid = 1
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/config/adom/" + adom + "/obj/vpnmgr/node",
                "data": [
                    {
                        "id": 0,
                        "protected_subnet": {
                            "addr": "all",
                            "seq": 1
                        },
                        "scope member": {
                            "name": devicename,
                            "vdom": "root"
                        },
                        "vpntable": overlayname,
                        "role": 0,
                        "iface": interface,
                        "hub_iface": [],
                        "peer": [],
                        "automatic_routing": 0,
                        "mode-cfg": 1,
                        "mode-cfg-ip-version": 0,
                        "ipv4-start-ip": str(oTemp[10]),
                        "ipv4-end-ip": str(oTemp[-1]),
                        "ipv4-netmask": str(oTemp.netmask),
                        "net-device": 0,
                        "tunnel-search": 1,
                        "extgwip": [],
                        "extgw_hubip": [],
                        "extgw_p2_per_net": 0,
                        "route-overlap": 0,
                        "vpn-zone": [],
                        "spoke-zone": [],
                        "vpn-interface-priority": 0,
                        "auto-configuration": 1,
                        "dns-service": 5,
                        # "dhcp-server": 1,
                        "ipsec-lease-hold": 60,
                        "add-route": 0,
                        "assign-ip": 1,
                        "assign-ip-from": 0,
                        "authusrgrp": [],
                        "dns-mode": 1,
                        "exchange-interface-ip": 0,
                        # "exchange-interface-ip": 0,
                        "peergrp": [],
                        "peertype": 1,
                        "unity-support": 1,
                        "xauthtype": 1
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("### add_vpn_hub ")
    json_addvpnhub = json.loads(res.text)
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    status_addvpnhub = json_addvpnhub['result'][0]['status']['message']
    return status_addvpnhub
